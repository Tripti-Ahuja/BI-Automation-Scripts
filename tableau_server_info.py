"""
Tableau Server Information Fetcher  (VDI-safe / IDLE-compatible)
=================================================================
Read-only script — only fetches metadata via Tableau REST API (GET requests).
Writes ONLY .xlsx output files to the same folder as this script.
No credentials are saved to disk.

Prerequisites (install via pip in IDLE or cmd):
    pip install tableauserverclient openpyxl truststore

Security notes:
    - Uses Personal Access Token (PAT) auth only — no passwords stored
    - PAT is held in memory only, never written to disk
    - SSL/TLS verification enabled via OS certificate store (truststore)
    - All API calls are read-only (GET); nothing is modified on the server
    - Session is signed out on exit, even if the script crashes
"""
import sys
import os
import re
import ssl
import time
import json
from datetime import datetime
from collections import defaultdict
from pathlib import Path

try:
    import tableauserverclient as TSC
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("")
    print("  Missing package: " + str(e))
    print("")
    print("  Install these in your VDI (run in cmd or IDLE terminal):")
    print("    pip install tableauserverclient openpyxl truststore")
    print("")
    input("  Press Enter to close...")
    sys.exit(1)


# ---------------------------------------------------------------------------
# SSL / TLS — use the OS certificate store (critical for corporate VDI)
# ---------------------------------------------------------------------------
_SSL_VERIFY = True
_SSL_STATUS = "unknown"

try:
    import truststore
    truststore.inject_into_ssl()
    _SSL_STATUS = "truststore (OS cert store)"
except ImportError:
    try:
        import certifi
        os.environ.setdefault("REQUESTS_CA_BUNDLE", certifi.where())
        _SSL_STATUS = "certifi bundle"
    except ImportError:
        _ctx = ssl.create_default_context()
        if _ctx.get_ca_certs():
            _SSL_STATUS = "Python default SSL context"
        else:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            _SSL_VERIFY = False
            _SSL_STATUS = "DISABLED (no cert store found)"


MAX_RETRIES = 3
RETRY_DELAY = 5


# ---------------------------------------------------------------------------
# URL Parsing
# ---------------------------------------------------------------------------
def parse_tableau_url(raw_url):
    """
    Accept a full browser URL like:
      https://us-east-1.online.tableau.com/#/site/mysite/explore
    and extract the base server URL and site ID automatically.
    """
    raw_url = raw_url.strip().rstrip("/")
    match = re.match(r'^(https?://[^/#]+)(?:/#/site/([^/]+))?', raw_url)
    if match:
        return match.group(1), match.group(2) or ""
    return raw_url, ""


# ---------------------------------------------------------------------------
# Authentication  — PAT only, nothing saved to disk
# ---------------------------------------------------------------------------
def _print_pat_instructions():
    print("  +---------------------------------------------------------+")
    print("  |  HOW TO CREATE A PERSONAL ACCESS TOKEN (one-time)       |")
    print("  +---------------------------------------------------------+")
    print("  |  1. Log into Tableau Cloud in your browser (SSO)        |")
    print("  |  2. Click your profile icon (top-right) -> My Account   |")
    print("  |  3. Under 'Personal Access Tokens' click + Create       |")
    print("  |  4. Give it a name (e.g. 'data-extractor')              |")
    print("  |  5. Copy the Token Name and Token Value                 |")
    print("  +---------------------------------------------------------+")
    print("")


def login():
    """Authenticate to Tableau using Personal Access Token (in-memory only)."""
    print("")
    print("  === TABLEAU SERVER DATA EXTRACTOR ===")
    print("  (read-only / VDI-safe)")
    print("")
    print("  SSL verification: " + _SSL_STATUS)
    if not _SSL_VERIFY:
        print("")
        print("  !! WARNING: SSL verification is OFF.")
        print("     Install truststore to fix:  pip install truststore")
        print("     Proceed only if you trust your network.")
        proceed = input("     Continue anyway? [y/N]: ").strip().lower()
        if proceed not in ("y", "yes"):
            print("  Aborted.")
            return None, None
    print("")

    raw_url = input("  Paste your Tableau URL (or base server URL): ").strip()
    base_url, detected_site = parse_tableau_url(raw_url)

    if detected_site:
        print("  -> Detected base URL : " + base_url)
        print("  -> Detected site ID  : " + detected_site)
        override = input("  Site ID [" + detected_site + "] (press Enter to keep): ").strip()
        site_id = override if override else detected_site
    else:
        print("  -> Using server URL  : " + base_url)
        site_id = input("  Site ID (leave blank for Default site): ").strip()

    print("")
    _print_pat_instructions()
    pat_name = input("  PAT Name : ").strip()
    pat_value = input("  PAT Value: ").strip()

    if not pat_name or not pat_value:
        print("  PAT Name and Value are both required.")
        return None, None

    auth = TSC.PersonalAccessTokenAuth(pat_name, pat_value, site_id=site_id)
    server = TSC.Server(base_url)
    server.add_http_options({"verify": _SSL_VERIFY})

    try:
        server.use_server_version()
    except Exception:
        server.version = "3.19"
        print("  (!) Could not auto-detect API version, using " + server.version)

    print("")
    print("  Signing in...")
    try:
        server.auth.sign_in(auth)
        print("  Success!  (API v" + server.version + ")")
        print("")
    except TSC.ServerResponseError as e:
        print("  FAILED: " + str(e.summary) + " - " + str(e.detail))
        return None, None
    except Exception as e:
        print("  FAILED: " + str(e))
        return None, None

    return server, auth


# ---------------------------------------------------------------------------
# Fetch helpers (with retry wrapper)
# ---------------------------------------------------------------------------
def _pager_to_list(server_endpoint, progress_label="items"):
    """Paginate a TSC endpoint with retry logic, return full list."""
    items = []
    for attempt in range(MAX_RETRIES):
        try:
            items = list(TSC.Pager(server_endpoint))
            break
        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                print("  [Retry " + str(attempt+1) + "/" + str(MAX_RETRIES) + ": " + str(e) + "]")
                time.sleep(RETRY_DELAY * (attempt + 1))
            else:
                print("  [Failed after " + str(MAX_RETRIES) + " attempts: " + str(e) + "]")
    return items


# ===================================================================
#  OPTION 1:  ALL PROJECTS
# ===================================================================
def fetch_projects(server):
    print("  [Getting projects...]")
    rows = []
    for p in _pager_to_list(server.projects, "projects"):
        rows.append({
            "ID": p.id,
            "Name": p.name,
            "Description": p.description or "",
            "Content Permissions": p.content_permissions or "",
            "Parent Project ID": p.parent_id or "",
            "Owner ID": p.owner_id or "",
        })
    print("  Done - " + str(len(rows)) + " projects")
    print("")
    return rows


# ===================================================================
#  OPTION 2:  ALL WORKBOOKS
# ===================================================================
def fetch_workbooks(server):
    print("  [Getting workbooks...]")
    rows = []
    for wb in _pager_to_list(server.workbooks, "workbooks"):
        rows.append({
            "ID": wb.id,
            "Name": wb.name,
            "Project Name": wb.project_name or "",
            "Project ID": wb.project_id or "",
            "Owner ID": wb.owner_id or "",
            "Content URL": wb.content_url or "",
            "Created At": str(wb.created_at or ""),
            "Updated At": str(wb.updated_at or ""),
            "Size (bytes)": wb.size or "",
            "Show Tabs": wb.show_tabs,
            "Tags": ", ".join(sorted(wb.tags)) if wb.tags else "",
        })
    print("  Done - " + str(len(rows)) + " workbooks")
    print("")
    return rows


# ===================================================================
#  OPTION 3:  ALL VIEWS
# ===================================================================
def fetch_views(server):
    print("  [Getting views...]")
    rows = []
    for v in _pager_to_list(server.views, "views"):
        rows.append({
            "ID": v.id,
            "Name": v.name,
            "Workbook ID": v.workbook_id or "",
            "Owner ID": v.owner_id or "",
            "Content URL": v.content_url or "",
            "Created At": str(v.created_at or ""),
            "Updated At": str(v.updated_at or ""),
            "Tags": ", ".join(sorted(v.tags)) if v.tags else "",
        })
    print("  Done - " + str(len(rows)) + " views")
    print("")
    return rows


# ===================================================================
#  OPTION 4:  ALL DATA SOURCES (metadata)
# ===================================================================
def fetch_datasources(server):
    print("  [Getting data sources...]")
    rows = []
    for ds in _pager_to_list(server.datasources, "data sources"):
        rows.append({
            "ID": ds.id,
            "Name": ds.name,
            "Type": ds.datasource_type or "",
            "Project Name": ds.project_name or "",
            "Project ID": ds.project_id or "",
            "Owner ID": ds.owner_id or "",
            "Content URL": ds.content_url or "",
            "Created At": str(ds.created_at or ""),
            "Updated At": str(ds.updated_at or ""),
            "Tags": ", ".join(sorted(ds.tags)) if ds.tags else "",
            "Has Extracts": getattr(ds, "has_extracts", ""),
            "Is Certified": getattr(ds, "certified", ""),
            "Certification Note": getattr(ds, "certification_note", ""),
        })
    print("  Done - " + str(len(rows)) + " data sources")
    print("")
    return rows


# ===================================================================
#  OPTION 5:  DATA SOURCE CONNECTIONS (deep-dive into each DS)
# ===================================================================
def fetch_datasource_connections(server):
    """Iterate every data source and pull its individual connection details."""
    all_ds = _pager_to_list(server.datasources, "data sources")
    total = len(all_ds)
    print("  [Fetching connections for " + str(total) + " data sources...]")

    rows = []
    for idx, ds in enumerate(all_ds, 1):
        if idx % 50 == 0 or idx == total:
            name_short = ds.name[:40] if len(ds.name) > 40 else ds.name
            print("    [" + str(idx) + "/" + str(total) + "] " + name_short)
        for attempt in range(MAX_RETRIES):
            try:
                server.datasources.populate_connections(ds)
                for conn in ds.connections:
                    rows.append({
                        "Data Source ID": ds.id,
                        "Data Source Name": ds.name,
                        "Connection ID": conn.id,
                        "Connection Type": conn.connection_type or "",
                        "Server Address": conn.server_address or "",
                        "Port": conn.server_port or "",
                        "Database Name": conn.datasource_name or "",
                        "Username": conn.username or "",
                        "Embed Password": conn.embed_password,
                    })
                break
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    rows.append({
                        "Data Source ID": ds.id,
                        "Data Source Name": ds.name,
                        "Connection ID": "ERROR",
                        "Connection Type": str(e),
                        "Server Address": "", "Port": "",
                        "Database Name": "", "Username": "",
                        "Embed Password": "",
                    })
        time.sleep(0.1)

    print("")
    print("  Done - " + str(len(rows)) + " connection entries")
    print("")
    return rows


# ===================================================================
#  OPTION 6:  ALL FLOWS
# ===================================================================
def fetch_flows(server):
    print("  [Getting flows...]")
    rows = []
    for f in _pager_to_list(server.flows, "flows"):
        rows.append({
            "ID": f.id,
            "Name": f.name,
            "Description": f.description or "",
            "Project Name": f.project_name or "",
            "Project ID": f.project_id or "",
            "Owner ID": f.owner_id or "",
            "Created At": str(f.created_at or ""),
            "Updated At": str(f.updated_at or ""),
            "Tags": ", ".join(sorted(f.tags)) if f.tags else "",
            "Web Page URL": f.webpage_url or "",
        })
    print("  Done - " + str(len(rows)) + " flows")
    print("")
    return rows


# ===================================================================
#  OPTION 7:  SERVER / SITE SUMMARY
# ===================================================================
def fetch_summary(server):
    print("  [Counting server objects...]")
    counts = defaultdict(int)
    for _ in _pager_to_list(server.projects):
        counts["Projects"] += 1
    for _ in _pager_to_list(server.workbooks):
        counts["Workbooks"] += 1
    for _ in _pager_to_list(server.views):
        counts["Views"] += 1
    for _ in _pager_to_list(server.datasources):
        counts["Data Sources"] += 1
    try:
        for _ in _pager_to_list(server.flows):
            counts["Flows"] += 1
    except Exception:
        counts["Flows"] = "N/A"

    rows = [{"Category": k, "Count": v} for k, v in counts.items()]

    print("")
    print("  +--------------------------+----------+")
    print("  |  Category                |  Count   |")
    print("  +--------------------------+----------+")
    for k, v in counts.items():
        print("  |  " + str(k).ljust(24) + " |  " + str(v).rjust(6) + "  |")
    print("  +--------------------------+----------+")
    print("")
    return rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="2F5496")
_HDR_ALIGN = Alignment(horizontal="center")


def _write_sheet(ws, rows):
    if not rows:
        ws["A1"] = "No data"
        return
    heads = list(rows[0].keys())
    for ci, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _HDR_ALIGN
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(heads, 1):
            v = row.get(h, "")
            ws.cell(row=ri, column=ci, value=str(v) if isinstance(v, (dict, list)) else v)
    for ci, h in enumerate(heads, 1):
        sample_lens = [len(str(row.get(h, ""))) for row in rows[:200]]
        max_len = max(len(h), max(sample_lens)) if sample_lens else len(h)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 14), 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _out_path(name):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "Tableau_" + name + "_" + ts + ".xlsx")


def _save_multi(sheets, filename):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        _write_sheet(ws, rows)
    path = _out_path(filename)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Interactive CLI
# ---------------------------------------------------------------------------
def main():
    server, auth = login()
    if not server:
        input("  Press Enter to close...")
        return

    try:
        while True:
            print("")
            print("  OPTIONS")
            print("  " + "-" * 50)
            print("  1. All Projects")
            print("  2. All Workbooks")
            print("  3. All Views")
            print("  4. All Data Sources")
            print("  5. Data Source Connections (deep-dive)")
            print("  6. All Flows")
            print("  7. Server / Site Summary (counts)")
            print("  8. Export ALL (everything in one file)")
            print("  0. Exit")

            ch = input("  Pick (0-8): ").strip()

            if ch == "0":
                break

            elif ch == "1":
                rows = fetch_projects(server)
                path = _save_multi({"Projects": rows}, "Projects")
                print("  Saved -> " + path + "  (" + str(len(rows)) + " rows)")
                input("  Press Enter to go back...")

            elif ch == "2":
                rows = fetch_workbooks(server)
                path = _save_multi({"Workbooks": rows}, "Workbooks")
                print("  Saved -> " + path + "  (" + str(len(rows)) + " rows)")
                input("  Press Enter to go back...")

            elif ch == "3":
                rows = fetch_views(server)
                path = _save_multi({"Views": rows}, "Views")
                print("  Saved -> " + path + "  (" + str(len(rows)) + " rows)")
                input("  Press Enter to go back...")

            elif ch == "4":
                rows = fetch_datasources(server)
                path = _save_multi({"Data Sources": rows}, "DataSources")
                print("  Saved -> " + path + "  (" + str(len(rows)) + " rows)")
                input("  Press Enter to go back...")

            elif ch == "5":
                ds_rows = fetch_datasources(server)
                conn_rows = fetch_datasource_connections(server)
                sheets = {"Data Sources": ds_rows, "Connections": conn_rows}
                path = _save_multi(sheets, "DataSource_Connections")
                print("  Saved -> " + path)
                print("    Data Sources:  " + str(len(ds_rows)) + " rows")
                print("    Connections:   " + str(len(conn_rows)) + " rows")
                input("  Press Enter to go back...")

            elif ch == "6":
                rows = fetch_flows(server)
                path = _save_multi({"Flows": rows}, "Flows")
                print("  Saved -> " + path + "  (" + str(len(rows)) + " rows)")
                input("  Press Enter to go back...")

            elif ch == "7":
                fetch_summary(server)
                input("  Press Enter to go back...")

            elif ch == "8":
                print("")
                print("  Fetching everything...")
                print("")
                print("  --- Projects ---")
                projects = fetch_projects(server)
                print("  --- Workbooks ---")
                workbooks = fetch_workbooks(server)
                print("  --- Views ---")
                views = fetch_views(server)
                print("  --- Data Sources ---")
                datasources = fetch_datasources(server)
                print("  --- Data Source Connections ---")
                connections = fetch_datasource_connections(server)
                print("  --- Flows ---")
                flows = fetch_flows(server)

                sheets = {
                    "Projects": projects,
                    "Workbooks": workbooks,
                    "Views": views,
                    "Data Sources": datasources,
                    "Connections": connections,
                    "Flows": flows,
                }
                path = _save_multi(sheets, "ALL")
                total = sum(len(v) for v in sheets.values())
                print("  Exported " + str(total) + " total rows -> " + path)
                print("    Projects:      " + str(len(projects)))
                print("    Workbooks:     " + str(len(workbooks)))
                print("    Views:         " + str(len(views)))
                print("    Data Sources:  " + str(len(datasources)))
                print("    Connections:   " + str(len(connections)))
                print("    Flows:         " + str(len(flows)))
                print("")
                print("  NOTE: file may contain sensitive data (server names, usernames)")
                input("  Press Enter to go back...")

            else:
                print("  Invalid choice.")

    finally:
        try:
            server.auth.sign_out()
            print("  Signed out from Tableau Server.")
        except Exception:
            pass

    print("  Bye!")
    print("")


if __name__ == "__main__":
    main()
