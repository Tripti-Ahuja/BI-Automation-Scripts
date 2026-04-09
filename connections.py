"""
Tableau Data Source Connection Extractor  (VDI-safe / IDLE-compatible)
======================================================================
Fetches ALL data source connections including embedded ones inside workbooks.
Read-only — only GET requests, writes only .xlsx output.

Prerequisites:  pip install tableauserverclient openpyxl truststore
"""
import sys
import os
import re
import ssl
import time
import json
from datetime import datetime
from pathlib import Path

try:
    import tableauserverclient as TSC
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("")
    print("  Missing package: " + str(e))
    print("  Install:  pip install tableauserverclient openpyxl truststore")
    print("")
    input("  Press Enter to close...")
    sys.exit(1)

# ---------------------------------------------------------------------------
# SSL / TLS
# ---------------------------------------------------------------------------
_SSL_VERIFY = True
_SSL_STATUS = "unknown"

try:
    import truststore
    truststore.inject_into_ssl()
    _SSL_STATUS = "truststore (OS cert store)"
except ImportError:
    try:
        import certifi
        os.environ.setdefault("REQUESTS_CA_BUNDLE", certifi.where())
        _SSL_STATUS = "certifi bundle"
    except ImportError:
        _ctx = ssl.create_default_context()
        if _ctx.get_ca_certs():
            _SSL_STATUS = "Python default SSL context"
        else:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            _SSL_VERIFY = False
            _SSL_STATUS = "DISABLED (no cert store found)"

MAX_RETRIES = 3
RETRY_DELAY = 5


# ---------------------------------------------------------------------------
# URL Parsing
# ---------------------------------------------------------------------------
def parse_tableau_url(raw_url):
    raw_url = raw_url.strip().rstrip("/")
    match = re.match(r'^(https?://[^/#]+)(?:/#/site/([^/]+))?', raw_url)
    if match:
        return match.group(1), match.group(2) or ""
    return raw_url, ""


# ---------------------------------------------------------------------------
# Authentication — PAT only, nothing saved to disk
# ---------------------------------------------------------------------------
def _print_pat_instructions():
    print("  +---------------------------------------------------------+")
    print("  |  HOW TO CREATE A PERSONAL ACCESS TOKEN (one-time)       |")
    print("  +---------------------------------------------------------+")
    print("  |  1. Log into Tableau Cloud in your browser (SSO)        |")
    print("  |  2. Click your profile icon (top-right) -> My Account   |")
    print("  |  3. Under 'Personal Access Tokens' click + Create       |")
    print("  |  4. Give it a name (e.g. 'data-extractor')              |")
    print("  |  5. Copy the Token Name and Token Value                 |")
    print("  +---------------------------------------------------------+")
    print("")


def login():
    print("")
    print("  === TABLEAU DATA SOURCE CONNECTION EXTRACTOR ===")
    print("  (includes embedded workbook connections)")
    print("")
    print("  SSL verification: " + _SSL_STATUS)
    if not _SSL_VERIFY:
        print("")
        print("  !! WARNING: SSL verification is OFF.")
        print("     Install truststore to fix:  pip install truststore")
        proceed = input("     Continue anyway? [y/N]: ").strip().lower()
        if proceed not in ("y", "yes"):
            return None
    print("")

    raw_url = input("  Paste your Tableau URL (or base server URL): ").strip()
    base_url, detected_site = parse_tableau_url(raw_url)

    if detected_site:
        print("  -> Detected base URL : " + base_url)
        print("  -> Detected site ID  : " + detected_site)
        override = input("  Site ID [" + detected_site + "] (press Enter to keep): ").strip()
        site_id = override if override else detected_site
    else:
        print("  -> Using server URL  : " + base_url)
        site_id = input("  Site ID (leave blank for Default site): ").strip()

    print("")
    _print_pat_instructions()
    pat_name = input("  PAT Name : ").strip()
    pat_value = input("  PAT Value: ").strip()

    if not pat_name or not pat_value:
        print("  PAT Name and Value are both required.")
        return None

    auth = TSC.PersonalAccessTokenAuth(pat_name, pat_value, site_id=site_id)
    server = TSC.Server(base_url)
    server.add_http_options({"verify": _SSL_VERIFY})

    try:
        server.use_server_version()
    except Exception:
        server.version = "3.19"
        print("  (!) Could not auto-detect API version, using " + server.version)

    print("")
    print("  Signing in...")
    try:
        server.auth.sign_in(auth)
        print("  Success!  (API v" + server.version + ")")
        print("")
    except TSC.ServerResponseError as e:
        print("  FAILED: " + str(e.summary) + " - " + str(e.detail))
        return None
    except Exception as e:
        print("  FAILED: " + str(e))
        return None

    return server


# ---------------------------------------------------------------------------
# Retry helper
# ---------------------------------------------------------------------------
def _pager_to_list(endpoint):
    for attempt in range(MAX_RETRIES):
        try:
            return list(TSC.Pager(endpoint))
        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                print("  [Retry " + str(attempt+1) + ": " + str(e) + "]")
                time.sleep(RETRY_DELAY * (attempt + 1))
            else:
                print("  [Failed: " + str(e) + "]")
    return []


# ---------------------------------------------------------------------------
# 1. Published data sources + their connections
# ---------------------------------------------------------------------------
def fetch_published_datasources(server):
    print("  [1/2] Getting PUBLISHED data sources...")
    all_ds = _pager_to_list(server.datasources)
    total = len(all_ds)
    print("  Found " + str(total) + " published data sources")

    ds_rows = []
    conn_rows = []

    for idx, ds in enumerate(all_ds, 1):
        if idx % 50 == 0 or idx == total or idx == 1:
            name_short = ds.name[:50] if len(ds.name) > 50 else ds.name
            print("    [" + str(idx) + "/" + str(total) + "] " + name_short)

        ds_rows.append({
            "ID": ds.id,
            "Name": ds.name,
            "Type": ds.datasource_type or "",
            "Source": "Published",
            "Project Name": ds.project_name or "",
            "Project ID": ds.project_id or "",
            "Owner ID": ds.owner_id or "",
            "Content URL": ds.content_url or "",
            "Created At": str(ds.created_at or ""),
            "Updated At": str(ds.updated_at or ""),
            "Tags": ", ".join(sorted(ds.tags)) if ds.tags else "",
            "Has Extracts": getattr(ds, "has_extracts", ""),
            "Is Certified": getattr(ds, "certified", ""),
            "Certification Note": getattr(ds, "certification_note", ""),
        })

        for attempt in range(MAX_RETRIES):
            try:
                server.datasources.populate_connections(ds)
                for conn in ds.connections:
                    conn_rows.append({
                        "Parent Type": "Published Data Source",
                        "Parent ID": ds.id,
                        "Parent Name": ds.name,
                        "Data Source Type": ds.datasource_type or "",
                        "Project Name": ds.project_name or "",
                        "Connection ID": conn.id,
                        "Connection Type": conn.connection_type or "",
                        "Server Address": conn.server_address or "",
                        "Port": conn.server_port or "",
                        "Database Name": conn.datasource_name or "",
                        "Username": conn.username or "",
                        "Embed Password": conn.embed_password,
                    })
                break
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    conn_rows.append({
                        "Parent Type": "Published Data Source",
                        "Parent ID": ds.id,
                        "Parent Name": ds.name,
                        "Data Source Type": ds.datasource_type or "",
                        "Project Name": ds.project_name or "",
                        "Connection ID": "ERROR",
                        "Connection Type": str(e),
                        "Server Address": "", "Port": "",
                        "Database Name": "", "Username": "",
                        "Embed Password": "",
                    })
        time.sleep(0.1)

    print("  Published: " + str(len(ds_rows)) + " data sources, "
          + str(len(conn_rows)) + " connections")
    print("")
    return ds_rows, conn_rows


# ---------------------------------------------------------------------------
# 2. Workbook-embedded connections (the ones missing from published list)
# ---------------------------------------------------------------------------
def fetch_workbook_connections(server):
    print("  [2/2] Getting WORKBOOK-EMBEDDED connections...")
    print("  (this captures data sources not published separately)")
    all_wb = _pager_to_list(server.workbooks)
    total = len(all_wb)
    print("  Found " + str(total) + " workbooks - scanning connections...")

    wb_conn_rows = []
    embedded_ds = {}

    for idx, wb in enumerate(all_wb, 1):
        if idx % 100 == 0 or idx == total:
            print("    [" + str(idx) + "/" + str(total) + "]")
        for attempt in range(MAX_RETRIES):
            try:
                server.workbooks.populate_connections(wb)
                for conn in wb.connections:
                    conn_type = conn.connection_type or ""
                    wb_conn_rows.append({
                        "Parent Type": "Workbook (embedded)",
                        "Parent ID": wb.id,
                        "Parent Name": wb.name,
                        "Data Source Type": conn_type,
                        "Project Name": wb.project_name or "",
                        "Connection ID": conn.id,
                        "Connection Type": conn_type,
                        "Server Address": conn.server_address or "",
                        "Port": conn.server_port or "",
                        "Database Name": conn.datasource_name or "",
                        "Username": conn.username or "",
                        "Embed Password": conn.embed_password,
                    })
                    key = (conn_type, conn.server_address or "", conn.datasource_name or "")
                    if key not in embedded_ds:
                        embedded_ds[key] = {
                            "ID": conn.id,
                            "Name": conn.datasource_name or conn_type,
                            "Type": conn_type,
                            "Source": "Embedded in Workbook",
                            "Project Name": wb.project_name or "",
                            "Project ID": wb.project_id or "",
                            "Owner ID": wb.owner_id or "",
                            "Content URL": "",
                            "Created At": "",
                            "Updated At": "",
                            "Tags": "",
                            "Has Extracts": "",
                            "Is Certified": "",
                            "Certification Note": "",
                            "Workbook(s)": wb.name,
                        }
                    else:
                        existing = embedded_ds[key].get("Workbook(s)", "")
                        if wb.name not in existing:
                            embedded_ds[key]["Workbook(s)"] = existing + ", " + wb.name
                break
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    wb_conn_rows.append({
                        "Parent Type": "Workbook (embedded)",
                        "Parent ID": wb.id,
                        "Parent Name": wb.name,
                        "Data Source Type": "ERROR",
                        "Project Name": wb.project_name or "",
                        "Connection ID": "ERROR",
                        "Connection Type": str(e),
                        "Server Address": "", "Port": "",
                        "Database Name": "", "Username": "",
                        "Embed Password": "",
                    })
        time.sleep(0.05)

    print("  Workbooks: " + str(len(wb_conn_rows)) + " embedded connections found")
    print("  Unique embedded data source signatures: " + str(len(embedded_ds)))
    print("")
    return wb_conn_rows, list(embedded_ds.values())


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="2F5496")
_HDR_ALIGN = Alignment(horizontal="center")


def _write_sheet(ws, rows):
    if not rows:
        ws["A1"] = "No data"
        return
    heads = list(rows[0].keys())
    for ci, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _HDR_ALIGN
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(heads, 1):
            v = row.get(h, "")
            ws.cell(row=ri, column=ci, value=str(v) if isinstance(v, (dict, list)) else v)
    for ci, h in enumerate(heads, 1):
        sample_lens = [len(str(row.get(h, ""))) for row in rows[:200]]
        max_len = max(len(h), max(sample_lens)) if sample_lens else len(h)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 14), 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_excel(sheets_dict):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets_dict.items():
        ws = wb.create_sheet(name[:31])
        _write_sheet(ws, rows)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "Tableau_Connections_" + ts + ".xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    server = login()
    if not server:
        input("  Press Enter to close...")
        return

    try:
        # Published data sources
        pub_ds_rows, pub_conn_rows = fetch_published_datasources(server)

        # Workbook-embedded connections
        wb_conn_rows, embedded_ds_rows = fetch_workbook_connections(server)

        # Merge all connections into one sheet
        all_connections = pub_conn_rows + wb_conn_rows

        # Merge all data sources (published + unique embedded)
        all_ds = pub_ds_rows.copy()
        pub_types_names = set()
        for r in pub_ds_rows:
            pub_types_names.add((r["Type"], r["Name"]))
        for emb in embedded_ds_rows:
            if (emb["Type"], emb["Name"]) not in pub_types_names:
                all_ds.append(emb)

        # Summary by connection type
        type_counts = {}
        for r in all_ds:
            t = r.get("Type", "unknown") or "unknown"
            type_counts[t] = type_counts.get(t, 0) + 1
        summary_rows = [{"Connection Type": k, "Count": v}
                        for k, v in sorted(type_counts.items(), key=lambda x: -x[1])]

        sheets = {
            "All Data Sources": all_ds,
            "All Connections": all_connections,
            "Published DS Only": pub_ds_rows,
            "Workbook Connections": wb_conn_rows,
            "Summary by Type": summary_rows,
        }

        path = save_excel(sheets)

        print("  " + "=" * 55)
        print("  RESULTS")
        print("  " + "=" * 55)
        print("  Published data sources:      " + str(len(pub_ds_rows)))
        print("  Embedded (workbook-only):     " + str(len(all_ds) - len(pub_ds_rows)))
        print("  TOTAL data sources:           " + str(len(all_ds)))
        print("  " + "-" * 55)
        print("  Published connections:        " + str(len(pub_conn_rows)))
        print("  Workbook connections:         " + str(len(wb_conn_rows)))
        print("  TOTAL connections:            " + str(len(all_connections)))
        print("  " + "=" * 55)
        print("")
        print("  Saved -> " + path)
        print("")
        print("  Sheets:")
        for name, rows in sheets.items():
            print("    " + name + ": " + str(len(rows)) + " rows")
        print("")
        print("  NOTE: file may contain sensitive data (server names, usernames)")

    finally:
        try:
            server.auth.sign_out()
            print("  Signed out.")
        except Exception:
            pass

    input("  Press Enter to close...")


if __name__ == "__main__":
    main()
