"""
Power BI Admin Data Extractor
Focus: Gateway Connections (DSN) & Workspace Details
Prerequisites: pip install msal requests openpyxl
     (VDI fix): pip install truststore
"""
import sys, os, re, time, json, threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import msal, requests
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\n  Run this first:  pip install msal requests openpyxl\n")
    sys.exit(1)

# ---------------------------------------------------------------------------
# SSL handling for corporate VDI / proxy environments
# ---------------------------------------------------------------------------
_SSL_VERIFY = True
_MSAL_HTTP_CLIENT = None

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    _SSL_VERIFY = False
    _s = requests.Session()
    _s.verify = False
    _MSAL_HTTP_CLIENT = _s

if not _SSL_VERIFY:
    print("  [!] truststore not found — SSL verification disabled.")
    print("      For the proper fix run:  pip install truststore\n")

API = "https://api.powerbi.com/v1.0/myorg"
ADM = API + "/admin"
CID = "ea0616ba-638b-4df5-95b9-636659ae5121"   # Microsoft Power BI PowerShell public client
SCOPES = ["https://analysis.windows.net/powerbi/api/.default"]
MAX_RETRIES = 3
RETRY_DELAY = 5
PARALLEL_WORKERS = 8


# ---------------------------------------------------------------------------
# Session — MSAL auth + automatic silent token refresh (thread-safe)
# ---------------------------------------------------------------------------
class Session:

    def __init__(self, app, account, token, expires_in):
        self._app = app
        self._account = account
        self._token = token
        self._expiry = time.time() + expires_in
        self._lock = threading.Lock()

    def token(self):
        with self._lock:
            if time.time() < self._expiry - 300:
                return self._token
            result = self._app.acquire_token_silent(SCOPES, account=self._account)
            if result and "access_token" in result:
                self._token = result["access_token"]
                self._expiry = time.time() + result.get("expires_in", 3600)
                print("  [Token refreshed]")
            return self._token


# ---------------------------------------------------------------------------
# Authentication
# ---------------------------------------------------------------------------
def login(tenant):
    kwargs = {}
    if _MSAL_HTTP_CLIENT is not None:
        kwargs["http_client"] = _MSAL_HTTP_CLIENT
    app = msal.PublicClientApplication(
        CID, authority="https://login.microsoftonline.com/" + tenant, **kwargs
    )
    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        print("  Login failed:", flow.get("error_description", "Unknown"))
        return None
    print("\n  1. Open:  https://microsoft.com/devicelogin")
    print("  2. Enter code:  " + flow["user_code"])
    print("  3. Sign in with your work email\n")
    print("  Waiting for sign-in...\n")
    r = app.acquire_token_by_device_flow(flow)
    if "access_token" in r:
        print("  Signed in!\n")
        accounts = app.get_accounts()
        return Session(app, accounts[0] if accounts else None,
                       r["access_token"], r.get("expires_in", 3600))
    print("  Sign-in failed:", r.get("error_description", "Unknown"))
    return None


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------
def get(session, url):
    """Paginated GET.  Returns list on success, None on first-page failure."""
    items, first = [], True
    while url:
        h = {"Authorization": "Bearer " + session.token()}
        ok = False
        for attempt in range(MAX_RETRIES):
            try:
                r = requests.get(url, headers=h, timeout=60, verify=_SSL_VERIFY)
            except Exception as e:
                print(f"  [Request error: {e}]")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                    continue
                return None if first else items
            if r.status_code == 200:
                d = r.json()
                items.extend(d.get("value", []))
                url = d.get("@odata.nextLink")
                ok = True
                break
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", RETRY_DELAY * (attempt + 1)))
                print(f"  [Rate limited — retrying in {wait}s]")
                time.sleep(wait)
                continue
            if r.status_code == 401:
                h = {"Authorization": "Bearer " + session.token()}
                continue
            return None if first else items
        if not ok:
            return None if first else items
        first = False
    return items


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _parse_conn(cd):
    if isinstance(cd, dict):
        return cd
    if isinstance(cd, str) and cd.strip():
        try:
            parsed = json.loads(cd)
            if isinstance(parsed, dict):
                return parsed
        except (json.JSONDecodeError, ValueError):
            return {"raw": cd}
    return {}


def _resolve_type(api_type, cd):
    if api_type != "Extension":
        return api_type
    for key in ("extensionDataSourceKind", "kind", "type"):
        val = cd.get(key, "")
        if val:
            return val
    path = cd.get("extensionDataSourcePath", "")
    if "snowflake" in path.lower():
        return "Snowflake"
    return "Extension"


def _resolve_server(cd, api_type):
    server = cd.get("server", "")
    if server:
        return server
    if api_type == "Extension":
        return cd.get("extensionDataSourcePath", cd.get("path", cd.get("raw", "")))
    return cd.get("path", cd.get("url", cd.get("raw", "")))


_ACCESS_LABELS = {
    "Read": "User",
    "ReadOverrideEffectiveIdentity": "User with resharing",
    "Owner": "Owner",
    "ReadWrite": "Owner",
    "Write": "Owner",
    "None": "None",
}


def _get_items(s, wid, item_type):
    """Fetch items (reports/datasets) from a workspace, trying admin then standard API."""
    items = get(s, ADM + f"/groups/{wid}/{item_type}")
    if items is not None:
        return items
    items = get(s, f"{API}/groups/{wid}/{item_type}")
    if items is not None:
        return items
    return []


def _get_ws_users(s, wid):
    users = get(s, ADM + f"/groups/{wid}/users?$top=500")
    if users:
        return users
    users = get(s, f"{API}/groups/{wid}/users")
    if users:
        return users
    return []


# ===================================================================
#  OPTION 1:  GATEWAY CONNECTIONS (DSN)
# ===================================================================
def fetch_gateway_data(s):
    print("  [Getting gateways...]")
    gw_list = get(s, ADM + "/gateways")
    if gw_list is None:
        gw_list = get(s, API + "/gateways") or []

    if not gw_list:
        print("  [No gateways found]")
        return [], []

    print(f"  [Found {len(gw_list)} gateway cluster(s)]\n")

    all_sources = []
    for gw in gw_list:
        gw_id = gw.get("id")
        if not gw_id:
            continue
        gw_cluster = gw.get("name", "")
        sources = get(s, ADM + f"/gateways/{gw_id}/datasources")
        if sources is None:
            sources = get(s, f"{API}/gateways/{gw_id}/datasources") or []
        print(f"  Gateway: {gw_cluster} — {len(sources)} connections")
        for src in sources:
            all_sources.append((gw_id, gw_cluster, src))

    if not all_sources:
        return [], []

    print(f"\n  [Fetching users for {len(all_sources)} datasources ({PARALLEL_WORKERS} threads)]")

    conn_rows = []
    user_rows = []
    lock = threading.Lock()
    counter = [0]

    def _process_source(args):
        gw_id, gw_cluster, src = args
        src_id = src.get("id", "")
        src_name = src.get("datasourceName", "")
        raw_type = src.get("datasourceType", "")
        cd = _parse_conn(src.get("connectionDetails", ""))
        conn_type = _resolve_type(raw_type, cd)
        server = _resolve_server(cd, raw_type)

        users = get(s, ADM + f"/gateways/{gw_id}/datasources/{src_id}/users")
        if users is None:
            users = get(s, f"{API}/gateways/{gw_id}/datasources/{src_id}/users")
        if users is None:
            users = []

        user_summary = ", ".join(
            u.get("displayName") or u.get("emailAddress", "")
            for u in users
        )

        c_row = {
            "Connection Name": src_name,
            "Connection Type": conn_type,
            "Server": server,
            "Database": cd.get("database", ""),
            "URL": cd.get("url", ""),
            "Path": cd.get("path", cd.get("extensionDataSourcePath", "")),
            "Users": user_summary,
            "# Users": len(users),
            "Gateway Cluster": gw_cluster,
            "Credential Type": src.get("credentialType", ""),
        }

        u_rows = []
        for u in users:
            raw = u.get("datasourceAccessRight", "")
            u_rows.append({
                "Connection Name": src_name,
                "Connection Type": conn_type,
                "Gateway Cluster": gw_cluster,
                "User Name": u.get("displayName", ""),
                "Email": u.get("emailAddress", ""),
                "Access (API)": _ACCESS_LABELS.get(raw, raw),
                "Actual Role": "Check Power BI UI",
                "Principal Type": u.get("principalType", ""),
            })

        with lock:
            conn_rows.append(c_row)
            user_rows.extend(u_rows)
            counter[0] += 1
            if counter[0] % 25 == 0 or counter[0] == len(all_sources):
                print(f"  [{counter[0]}/{len(all_sources)}] datasources", flush=True)

    with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
        futs = [pool.submit(_process_source, a) for a in all_sources]
        for f in as_completed(futs):
            try:
                f.result()
            except Exception as e:
                print(f"  [Worker error: {e}]")

    print(f"\n  Done — {len(conn_rows)} connections, {len(user_rows)} user entries\n")
    return conn_rows, user_rows


# ===================================================================
#  OPTION 2:  WORKSPACE DETAILS
# ===================================================================
def fetch_workspace_data(s):
    print("  [Getting workspace list...]")
    ws_list = get(s, ADM + "/groups?$top=5000")
    if ws_list is None:
        ws_list = get(s, API + "/groups") or []

    if not ws_list:
        print("  [No workspaces found]")
        return [], []

    total = len(ws_list)
    print(f"  [Found {total} workspaces — fetching details ({PARALLEL_WORKERS} threads)]")
    print(f"  [This will take a few minutes...]\n")

    overview_rows = []
    access_rows = []
    lock = threading.Lock()
    counter = [0]

    def _process_workspace(ws):
        try:
            wid = ws.get("id", "")
            ws_name = ws.get("name", "")

            users = _get_ws_users(s, wid)
            reports = _get_items(s, wid, "reports")
            datasets = _get_items(s, wid, "datasets")

            owners = [
                u.get("displayName") or u.get("emailAddress", "")
                for u in users
                if u.get("groupUserAccessRight") == "Admin"
            ]

            o_row = {
                "Workspace Name": ws_name,
                "Workspace ID": wid,
                "Type": ws.get("type", ""),
                "State": ws.get("state", ""),
                "Owner(s)": ", ".join(owners),
                "# Reports": len(reports),
                "# Semantic Models": len(datasets),
                "# Users/Groups with Access": len(users),
                "Capacity ID": ws.get("capacityId", ""),
            }

            a_rows = []
            for u in users:
                a_rows.append({
                    "Workspace": ws_name,
                    "Workspace ID": wid,
                    "User / Group Name": u.get("displayName", ""),
                    "Email": u.get("emailAddress", ""),
                    "Role": u.get("groupUserAccessRight", ""),
                    "Principal Type": u.get("principalType", ""),
                    "Identifier (UPN)": u.get("identifier", ""),
                })

            with lock:
                overview_rows.append(o_row)
                access_rows.extend(a_rows)
                counter[0] += 1
                if counter[0] % 50 == 0 or counter[0] == total:
                    print(f"  [{counter[0]}/{total}] workspaces", flush=True)
        except Exception as e:
            with lock:
                counter[0] += 1
            print(f"  [Error: {ws.get('name', '?')}: {e}]")

    with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
        list(pool.map(_process_workspace, ws_list))

    print(f"\n  Done — {len(overview_rows)} workspaces, {len(access_rows)} access entries\n")
    return overview_rows, access_rows


# ===================================================================
#  Gateway lookup for DSN mapping
# ===================================================================
def _build_gateway_lookup(s):
    """Build lookup: datasource_id -> {Connection Name, Gateway Cluster}."""
    lookup = {}
    gw_list = get(s, ADM + "/gateways")
    if gw_list is None:
        gw_list = get(s, API + "/gateways") or []
    for gw in gw_list:
        gw_id = gw.get("id")
        if not gw_id:
            continue
        gw_cluster = gw.get("name", "")
        sources = get(s, ADM + f"/gateways/{gw_id}/datasources")
        if sources is None:
            sources = get(s, f"{API}/gateways/{gw_id}/datasources") or []
        for src in sources:
            src_id = src.get("id", "")
            if src_id:
                lookup[src_id] = {
                    "Connection Name": src.get("datasourceName", ""),
                    "Gateway Cluster": gw_cluster,
                }
    return lookup


# ===================================================================
#  OPTION 3:  WORKSPACE REPORTS & SEMANTIC MODELS (detailed)
#
#  Uses tenant-wide admin endpoints for reports & datasets which
#  return modifiedDateTime (the per-workspace endpoint does not).
#  Then fetches refresh dates and DSN mapping in parallel phases.
# ===================================================================
def fetch_workspace_items(s):
    """Fetch all reports and semantic models with dates and DSN mapping."""

    # ---- Gateway lookup for DSN mapping ----
    print("  [Building gateway connection lookup...]")
    gw_lookup = _build_gateway_lookup(s)
    print(f"  [Indexed {len(gw_lookup)} gateway data sources]\n")

    # ---- Workspace name map ----
    print("  [Getting workspace list...]")
    ws_list = get(s, ADM + "/groups?$top=5000")
    if ws_list is None:
        ws_list = get(s, API + "/groups") or []
    ws_name_map = {ws.get("id", ""): ws.get("name", "") for ws in ws_list}
    print(f"  [Found {len(ws_list)} workspaces]\n")

    # ==================================================================
    #  PHASE 1 — Tenant-wide fetch of ALL reports and datasets
    #  /admin/reports  and  /admin/datasets  return modifiedDateTime
    # ==================================================================
    print("  Phase 1/3 — fetching all reports & datasets (tenant-wide)...")

    all_reports = get(s, ADM + "/reports?$top=5000")
    if all_reports is None:
        all_reports = []
    print(f"    Reports fetched: {len(all_reports)}")

    all_datasets = get(s, ADM + "/datasets?$top=5000")
    if all_datasets is None:
        all_datasets = []
    print(f"    Datasets fetched: {len(all_datasets)}\n")

    # ==================================================================
    #  PHASE 2 — Refresh dates for refreshable datasets (parallel)
    # ==================================================================
    refreshable = [
        (ds.get("workspaceId", ""), ds.get("id", ""))
        for ds in all_datasets
        if ds.get("isRefreshable")
    ]

    refresh_map = {}
    lock = threading.Lock()

    if refreshable:
        print(f"  Phase 2/3 — fetching refresh dates for {len(refreshable)} datasets ({PARALLEL_WORKERS} threads)")
        counter = [0]
        ref_total = len(refreshable)

        def _fetch_refresh(pair):
            wid, did = pair
            last_refresh = ""
            try:
                refreshes = get(s, f"{API}/groups/{wid}/datasets/{did}/refreshes?$top=1")
                if refreshes:
                    last_refresh = refreshes[0].get("endTime",
                                                     refreshes[0].get("startTime", ""))
            except Exception:
                pass
            with lock:
                refresh_map[(wid, did)] = last_refresh
                counter[0] += 1
                c = counter[0]
                if c % 100 == 0 or c == ref_total:
                    print(f"    [{c}/{ref_total}] refresh dates", flush=True)

        with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
            list(pool.map(_fetch_refresh, refreshable))
        print()
    else:
        print("  Phase 2/3 — no refreshable datasets, skipping\n")

    # ==================================================================
    #  PHASE 3 — DSN mapping for gateway-connected datasets (parallel)
    # ==================================================================
    gw_datasets = [
        (ds.get("workspaceId", ""), ds.get("id", ""))
        for ds in all_datasets
        if ds.get("isOnPremGatewayRequired") and gw_lookup
    ]

    dsn_map = {}

    if gw_datasets:
        print(f"  Phase 3/3 — mapping DSN connections for {len(gw_datasets)} datasets ({PARALLEL_WORKERS} threads)")
        counter = [0]
        dsn_total = len(gw_datasets)

        def _fetch_dsn(pair):
            wid, did = pair
            names, clusters = [], []
            try:
                ds_sources = get(s, f"{API}/groups/{wid}/datasets/{did}/datasources")
                if ds_sources:
                    for dss in ds_sources:
                        gw_ds_id = dss.get("datasourceId", "")
                        if gw_ds_id and gw_ds_id in gw_lookup:
                            names.append(gw_lookup[gw_ds_id]["Connection Name"])
                            clusters.append(gw_lookup[gw_ds_id]["Gateway Cluster"])
            except Exception:
                pass
            with lock:
                dsn_map[(wid, did)] = {
                    "Connection Name": ", ".join(dict.fromkeys(names)),
                    "Gateway Cluster": ", ".join(dict.fromkeys(clusters)),
                }
                counter[0] += 1
                c = counter[0]
                if c % 50 == 0 or c == dsn_total:
                    print(f"    [{c}/{dsn_total}] DSN lookups", flush=True)

        with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as pool:
            list(pool.map(_fetch_dsn, gw_datasets))
        print()
    else:
        print("  Phase 3/3 — no gateway datasets to map, skipping\n")

    # ==================================================================
    #  Assemble output rows
    # ==================================================================
    rows = []

    for rpt in all_reports:
        wid = rpt.get("workspaceId", "")
        rows.append({
            "Workspace Name": ws_name_map.get(wid, wid),
            "Workspace ID": wid,
            "Item Type": "Report",
            "Name": rpt.get("name", ""),
            "ID": rpt.get("id", ""),
            "Created Date": rpt.get("createdDateTime", ""),
            "Modified Date": rpt.get("modifiedDateTime", ""),
            "Last Refresh Date": "",
            "DSN Connection Name": "",
            "Gateway Cluster": "",
        })

    for ds in all_datasets:
        wid = ds.get("workspaceId", "")
        did = ds.get("id", "")
        dsn_info = dsn_map.get((wid, did), {})
        rows.append({
            "Workspace Name": ws_name_map.get(wid, wid),
            "Workspace ID": wid,
            "Item Type": "Semantic Model",
            "Name": ds.get("name", ""),
            "ID": did,
            "Created Date": ds.get("createdDate", ""),
            "Modified Date": ds.get("modifiedDateTime", ds.get("configuredBy", "")),
            "Last Refresh Date": refresh_map.get((wid, did), ""),
            "DSN Connection Name": dsn_info.get("Connection Name", ""),
            "Gateway Cluster": dsn_info.get("Gateway Cluster", ""),
        })

    rpt_count = sum(1 for r in rows if r["Item Type"] == "Report")
    model_count = sum(1 for r in rows if r["Item Type"] == "Semantic Model")
    print(f"  Done — {rpt_count} reports, {model_count} semantic models\n")
    return rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="2F5496")
_HDR_ALIGN = Alignment(horizontal="center")


def _write_sheet(ws, rows):
    if not rows:
        ws["A1"] = "No data"
        return
    heads = list(rows[0].keys())
    for ci, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, _HDR_ALIGN
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(heads, 1):
            v = row.get(h, "")
            ws.cell(row=ri, column=ci, value=str(v) if isinstance(v, (dict, list)) else v)
    for ci, h in enumerate(heads, 1):
        sample_lens = [len(str(row.get(h, ""))) for row in rows[:200]]
        max_len = max(len(h), max(sample_lens)) if sample_lens else len(h)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 14), 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _out_path(name):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        f"PowerBI_{name}_{ts}.xlsx")


def _save_multi(sheets, filename):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        _write_sheet(ws, rows)
    path = _out_path(filename)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Interactive CLI
# ---------------------------------------------------------------------------
def main():
    print("\n  === POWER BI DATA EXTRACTOR ===\n")
    url = input("  Paste your Power BI URL (or press Enter): ").strip() or "app.powerbi.com"
    m = re.search(r"ctid=([a-f0-9\-]{36})", url, re.I)
    tenant = m.group(1) if m else "common"

    session = login(tenant)
    if not session:
        input("\n  Press Enter to close...")
        return

    while True:
        print("\n  OPTIONS")
        print("  " + "-" * 35)
        print("  1. Gateway Connections (DSN)")
        print("  2. Workspace Details")
        print("  3. Reports & Semantic Models (detailed)")
        print("  4. Export ALL (everything in one file)")
        print("  0. Exit")

        ch = input("\n  Pick (0-4): ").strip()

        if ch == "0":
            print("\n  Bye!\n")
            break

        if ch == "1":
            print("\n  Fetching Gateway Connections...\n")
            conns, users = fetch_gateway_data(session)
            sheets = {"Connections": conns, "Connection Users": users}
            path = _save_multi(sheets, "Gateway_Connections")
            print(f"  Saved -> {path}")
            print(f"  Sheets:  Connections ({len(conns)} rows)  |  "
                  f"Connection Users ({len(users)} rows)")
            input("\n  Press Enter to go back...")

        elif ch == "2":
            print("\n  Fetching Workspace Details...\n")
            overview, access = fetch_workspace_data(session)
            sheets = {"Workspace Overview": overview, "Workspace Access": access}
            path = _save_multi(sheets, "Workspace_Details")
            print(f"  Saved -> {path}")
            print(f"  Sheets:  Workspace Overview ({len(overview)} rows)  |  "
                  f"Workspace Access ({len(access)} rows)")
            input("\n  Press Enter to go back...")

        elif ch == "3":
            print("\n  Fetching Reports & Semantic Models...\n")
            ws_items = fetch_workspace_items(session)
            sheets = {"Reports & Semantic Models": ws_items}
            path = _save_multi(sheets, "Reports_SemanticModels")
            print(f"  Saved -> {path}")
            rpt_n = sum(1 for r in ws_items if r.get("Item Type") == "Report")
            sm_n = sum(1 for r in ws_items if r.get("Item Type") == "Semantic Model")
            print(f"  Reports: {rpt_n}  |  Semantic Models: {sm_n}")
            input("\n  Press Enter to go back...")

        elif ch == "4":
            print("\n  Fetching everything...\n")
            print("  --- Gateway Connections ---\n")
            conns, conn_users = fetch_gateway_data(session)
            print("  --- Workspace Details ---\n")
            overview, ws_access = fetch_workspace_data(session)
            print("  --- Reports & Semantic Models ---\n")
            ws_items = fetch_workspace_items(session)

            sheets = {
                "Connections": conns,
                "Connection Users": conn_users,
                "Workspace Overview": overview,
                "Workspace Access": ws_access,
                "Reports & Semantic Models": ws_items,
            }
            path = _save_multi(sheets, "ALL")
            total = sum(len(v) for v in sheets.values())
            print(f"  Exported {total} total rows -> {path}")
            print(f"    Connections:              {len(conns)}")
            print(f"    Connection Users:         {len(conn_users)}")
            print(f"    Workspace Overview:       {len(overview)}")
            print(f"    Workspace Access:         {len(ws_access)}")
            print(f"    Reports & Semantic Models: {len(ws_items)}")
            print("\n  NOTE: file contains sensitive data (server names, emails, etc.)")
            input("\n  Press Enter to go back...")

        else:
            print("  Invalid choice.")


if __name__ == "__main__":
    main()
